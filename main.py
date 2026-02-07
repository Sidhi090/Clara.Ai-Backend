import pandas as pd
from pathlib import Path
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================
BASE_DIR = Path("IRC")

ERP_FILE = BASE_DIR / "erp data.xlsx"
TALLY_FILE = BASE_DIR / "tally dump till date for reco 31.10.2025 (2).xlsx"

GST_FILES = {
    "0.5": BASE_DIR / "05_ CGST (1).xls",
    "2.5": BASE_DIR / "2.5_ CGST (1).xls",
    "9": BASE_DIR / "9_ CGST (1).xls",
    "6": None  # provision for future
}

REG_AGREEMENT_FILE = BASE_DIR / "registered agreement data.xls"

# =========================================================
# STEP 0 – LOAD ERP DATA & CREATE UNIQUE IDENTIFIER
# =========================================================
print("="*70)
print("STEP 0: Loading ERP Data & Creating Unique Identifiers")
print("="*70)

raw_df = pd.read_excel(ERP_FILE, header=None)

# Normalize all cells to string for safe scanning
raw_df = raw_df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")

# ------------------------------------------------------
# STEP 2: Define required columns (logical names)
# ------------------------------------------------------
REQUIRED_COLUMNS = {
    "member_name": ["MEMBER NAME"],
    "unit_code": ["UNIT CODE"],
    "total_received": [
        "TOTAL RECEIVED AMOUNT",
    ],
}

# ------------------------------------------------------
# STEP 3: Find the header row dynamically
# ------------------------------------------------------
print("\n[STEP 3] Scanning rows to detect header...")

header_row_index = None
header_map = {}

for row_idx in range(len(raw_df)):
    row_values = raw_df.iloc[row_idx].str.upper().tolist()

    print(f"\n🔍 Scanning Excel Row {row_idx + 1}")

    if row_idx == 10:
        break
    print(f"Row values: {row_values}")

    temp_map = {}

    for logical_name, possible_names in REQUIRED_COLUMNS.items():
        for col_idx, cell in enumerate(row_values):
            for name in possible_names:
                if name in cell and cell != "":
                    temp_map[logical_name] = col_idx
                    print(
                        f"   ✅ Found '{name}' for '{logical_name}' "
                        f"at column index {col_idx}"
                    )

    print(f"   ➜ Matched columns so far: {temp_map}")

    if len(temp_map) == len(REQUIRED_COLUMNS):
        header_row_index = row_idx
        header_map = temp_map
        print(f"\n🎯 HEADER ROW CONFIRMED at Excel row {row_idx + 1}")
        break

if header_row_index is None:
    raise ValueError("❌ Header row not found. Required columns missing.")

# ------------------------------------------------------
# STEP 4: Load ERP using detected header
# ------------------------------------------------------
print("\n[STEP 4] Reloading Excel using detected header row...")
erp_df = pd.read_excel(ERP_FILE, header=header_row_index)

print(f"ERP DataFrame shape: {erp_df.shape}")
print("Detected column names:")
for i, col in enumerate(erp_df.columns):
    print(f"  {i}: {col}")

# Clean column names
erp_df.columns = erp_df.columns.astype(str).str.strip()

# Drop empty rows
before_rows = len(erp_df)
erp_df = erp_df.dropna(how="all")
after_rows = len(erp_df)

print(f"Dropped {before_rows - after_rows} empty rows")

# ------------------------------------------------------
# STEP 5: Resolve actual column names
# ------------------------------------------------------
member_name_col = erp_df.columns[header_map["member_name"]]
unit_code_col = erp_df.columns[header_map["unit_code"]]
total_received_col = erp_df.columns[header_map["total_received"]]

print("\n[STEP 5] Final column mapping:")
print(f"  Member Name Column        → {member_name_col}")
print(f"  Unit Code Column          → {unit_code_col}")
print(f"  Total Received Amount Col → {total_received_col}")

# ------------------------------------------------------
# STEP 6: Create UniqueID
# ------------------------------------------------------
print("\n[STEP 6] Creating UniqueID...")

erp_df["UniqueID"] = (
    erp_df[unit_code_col].astype(str).str.strip()
    + " "
    + erp_df[member_name_col].astype(str).str.strip()
)

# Remove junk rows
before_rows = len(erp_df)
erp_df = erp_df[
    (erp_df[unit_code_col].astype(str).str.strip() != "")
    & (erp_df[member_name_col].astype(str).str.strip() != "")
]
after_rows = len(erp_df)

print(f"Removed {before_rows - after_rows} invalid rows")

print("\nSample rows with UniqueID:")
print(
    erp_df[
        [unit_code_col, member_name_col, total_received_col, "UniqueID"]
    ].head(10)
)

print("\n✅ DONE: ERP data processed successfully")


# Ensure Total Received Amount is numeric
if total_received_col:
    erp_df["Total_Received_Amount"] = pd.to_numeric(
        erp_df[total_received_col].astype(str).str.replace(',', ''), 
        errors='coerce'
    ).fillna(0)
else:
    erp_df["Total_Received_Amount"] = 0

erp_unique_ids = set(erp_df["UniqueID"])

print(f"\n✓ Loaded {len(erp_df)} records from ERP")
print(f"✓ Created {len(erp_unique_ids)} unique identifiers")


# =========================================================
# STEP 1A – LOAD TALLY DATA
# =========================================================
print("\n" + "="*70)
print("STEP 1A: Loading Tally Data")
print("="*70)

print("\n[STEP 1] Loading entire Tally sheet without header...")

raw_df = pd.read_excel(TALLY_FILE, header=None)
raw_df = raw_df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")

print(f"Loaded raw sheet with shape: {raw_df.shape}")

print("\nPreview first 15 rows:")
print(raw_df.head(15))

# ------------------------------------------------------
# STEP 2: Required columns
# ------------------------------------------------------
print("\n[STEP 2] Required columns definition...")

REQUIRED_COLUMNS = {
    "particulars": ["PARTICULARS"],
    "debit": ["DEBIT"],
    "credit": ["CREDIT"],
}

for k, v in REQUIRED_COLUMNS.items():
    print(f"  {k}: {v}")

# ------------------------------------------------------
# STEP 3: Detect header row
# ------------------------------------------------------
print("\n[STEP 3] Detecting MULTI-ROW header...")

header_row_index = None
header_map = {}

WINDOW_SIZE = 4  # how many consecutive rows to inspect

for start_row in range(len(raw_df) - WINDOW_SIZE):
    print(f"\n🔍 Checking rows {start_row + 1} to {start_row + WINDOW_SIZE}")

    combined_cells = {}

    for row_offset in range(WINDOW_SIZE):
        row_idx = start_row + row_offset
        row_values = raw_df.iloc[row_idx].str.upper().tolist()

        print(f"   Row {row_idx + 1}: {row_values}")

        for col_idx, cell in enumerate(row_values):
            if cell == "":
                continue

            # Collect all header words seen per column
            combined_cells.setdefault(col_idx, set()).add(cell)

    temp_map = {}

    for logical_col, possible_names in REQUIRED_COLUMNS.items():
        for col_idx, texts in combined_cells.items():
            for text in texts:
                for name in possible_names:
                    if name in text:
                        temp_map[logical_col] = col_idx
                        print(
                            f"   ✅ Found '{name}' for '{logical_col}' "
                            f"in column {col_idx}"
                        )

    print(f"   ➜ Combined matches: {temp_map}")

    if len(temp_map) == len(REQUIRED_COLUMNS):
        header_row_index = start_row
        header_map = temp_map
        print(
            f"\n🎯 HEADER CONFIRMED using rows "
            f"{start_row + 1} to {start_row + WINDOW_SIZE}"
        )
        break

if header_row_index is None:
    raise ValueError("❌ Header row not detected even with multi-row logic")


# ------------------------------------------------------
# STEP 4: Reload using detected header
# ------------------------------------------------------
print("\n[STEP 4] Reloading Excel using detected header...")

tally_df = pd.read_excel(
    TALLY_FILE,
    header=header_row_index
)

tally_df.columns = tally_df.columns.astype(str).str.strip()
tally_df = tally_df.dropna(how="all")

print(f"Tally DataFrame shape after reload: {tally_df.shape}")

print("\nDetected columns:")
for i, col in enumerate(tally_df.columns):
    print(f"  {i}: {col}")

# ------------------------------------------------------
# STEP 5: Resolve column names
# ------------------------------------------------------
particulars_col = tally_df.columns[header_map["particulars"]]
debit_col = tally_df.columns[header_map["debit"]]
credit_col = tally_df.columns[header_map["credit"]]

print("\n[STEP 5] Column mapping:")
print(f"  Particulars → {particulars_col}")
print(f"  Debit       → {debit_col}")
print(f"  Credit      → {credit_col}")

# ------------------------------------------------------
# STEP 6: Find 'Flats' row and cut data
# ------------------------------------------------------
print("\n[STEP 6] Locating 'Flats' row...")

flats_mask = tally_df[particulars_col].str.upper() == "FLATS"

if not flats_mask.any():
    raise ValueError("❌ 'Flats' row not found — cannot determine data start")

flats_row_index = flats_mask.idxmax()

print(f"✅ 'Flats' found at DataFrame index: {flats_row_index}")

print("\nRows ABOVE Flats (to be removed):")
print(tally_df.loc[:flats_row_index, [particulars_col]].tail(10))

# Keep only rows AFTER Flats
tally_df = tally_df.loc[flats_row_index + 1:].reset_index(drop=True)

print(f"\nRows remaining AFTER Flats: {len(tally_df)}")

# ------------------------------------------------------
# STEP 7: Create UniqueID
# ------------------------------------------------------
print("\n[STEP 7] Creating UniqueID...")

tally_df["UniqueID"] = tally_df[particulars_col].astype(str).str.strip()

# ------------------------------------------------------
# STEP 8: Convert Debit / Credit
# ------------------------------------------------------
print("\n[STEP 8] Converting Debit & Credit to numeric...")

tally_df["Debit"] = pd.to_numeric(
    tally_df[debit_col].astype(str).str.replace(",", ""),
    errors="coerce"
).fillna(0)

tally_df["Credit"] = pd.to_numeric(
    tally_df[credit_col].astype(str).str.replace(",", ""),
    errors="coerce"
).fillna(0)

# ------------------------------------------------------
# STEP 9: Remove junk rows after Flats
# ------------------------------------------------------
print("\n[STEP 9] Removing non-flat rows...")

before = len(tally_df)

tally_df = tally_df[
    tally_df["UniqueID"].str.contains("SM-", case=False, na=False)
]

after = len(tally_df)

print(f"Removed {before - after} non-flat rows")

# ------------------------------------------------------
# STEP 10: Final check
# ------------------------------------------------------
print("\n[STEP 10] Final data preview:")
print(
    tally_df[
        ["UniqueID", "Debit", "Credit"]
    ].head(10)
)

print(f"\n✅ FINAL TALLY RECORD COUNT: {len(tally_df)}")


# =========================================================
# STEP 1B – COMPARE UNIQUE IDs (ERP vs TALLY)
# =========================================================
print("\n" + "=" * 80)
print("STEP 1B: Comparing Unique Identifiers (CLEAN + NORMALIZED)")
print("=" * 80)

# ---------------------------------------------------------
# STEP 1: Normalization function
# ---------------------------------------------------------
def normalize_unique_id(val: str) -> str:
    """
    Normalize UniqueID for reliable comparison
    """
    if val is None:
        return ""

    val = str(val).upper().strip()

    # Remove multiple spaces
    while "  " in val:
        val = val.replace("  ", " ")

    # Remove trailing Dr / Cr
    if val.endswith(" DR"):
        val = val[:-3]
    if val.endswith(" CR"):
        val = val[:-3]

    # Remove pure NAN values
    if val in {"NAN", "NAN NAN"}:
        return ""

    return val


# ---------------------------------------------------------
# STEP 2: Normalize & CLEAN ERP UniqueIDs
# ---------------------------------------------------------
print("\n[STEP 2] Normalizing ERP UniqueIDs...")

erp_df["NormalizedUniqueID"] = erp_df["UniqueID"].apply(normalize_unique_id)

# 🔥 REMOVE invalid ERP IDs
erp_df = erp_df[
    erp_df["NormalizedUniqueID"].notna()
    & (erp_df["NormalizedUniqueID"] != "")
]

erp_unique_ids = set(erp_df["NormalizedUniqueID"])

print(f"✓ ERP UniqueIDs count (after cleanup): {len(erp_unique_ids)}")

# ---------------------------------------------------------
# STEP 3: Normalize & CLEAN Tally UniqueIDs
# ---------------------------------------------------------
print("\n[STEP 3] Normalizing Tally UniqueIDs...")

tally_df["NormalizedUniqueID"] = tally_df["UniqueID"].apply(normalize_unique_id)

# 🔥 REMOVE invalid Tally IDs
tally_df = tally_df[
    tally_df["NormalizedUniqueID"].notna()
    & (tally_df["NormalizedUniqueID"] != "")
]

tally_unique_ids = set(tally_df["NormalizedUniqueID"])

print(f"✓ Tally UniqueIDs count (after cleanup): {len(tally_unique_ids)}")

# ---------------------------------------------------------
# STEP 4: Compare sets
# ---------------------------------------------------------
print("\n[STEP 4] Comparing ERP vs Tally UniqueIDs...")

in_erp_not_tally = erp_unique_ids - tally_unique_ids
in_tally_not_erp = tally_unique_ids - erp_unique_ids
common_customers = erp_unique_ids & tally_unique_ids

print("\n📊 Comparison Results:")
print(f"  ✓ Common customers              : {len(common_customers)}")
print(f"  ⚠️  In ERP but NOT in Tally     : {len(in_erp_not_tally)}")
print(f"  ⚠️  In Tally but NOT in ERP     : {len(in_tally_not_erp)}")

# ---------------------------------------------------------
# STEP 5: Debug output
# ---------------------------------------------------------
if in_erp_not_tally:
    print("\n⚠️ Customers in ERP but missing in Tally:")
    for customer in sorted(list(in_erp_not_tally))[:10]:
        print(f"  - {customer}")
    if len(in_erp_not_tally) > 10:
        print(f"  ... and {len(in_erp_not_tally) - 10} more")

if in_tally_not_erp:
    print("\n⚠️ Customers in Tally but missing in ERP:")
    for customer in sorted(list(in_tally_not_erp))[:10]:
        print(f"  - {customer}")
    if len(in_tally_not_erp) > 10:
        print(f"  ... and {len(in_tally_not_erp) - 10} more")

# ---------------------------------------------------------
# STEP 6: Sanity check
# ---------------------------------------------------------
print("\n[STEP 6] Sample matched customers:")
for uid in list(common_customers)[:5]:
    print(f"  ✓ MATCHED: {uid}")

print("\n✅ STEP 1B COMPLETED – NAN ISSUE FIXED")


# =========================================================
# STEP 1C – NET BALANCE CALCULATION
# =========================================================
print("\n" + "="*70)
print("STEP 1C: Calculating Net Balance (Credit - Debit)")
print("="*70)

# Net Balance = Credit - Debit (shows as negative Dr)
tally_df["net_balance"] = tally_df["Credit"] - tally_df["Debit"]

print(f"\n✓ Net Balance calculated for {len(tally_df)} records")


# =========================================================
# STEP 2 – PERMANENT DIFFERENCE (GST + REGISTERED AGREEMENT)
# =========================================================

print("\n" + "="*70)
print("STEP 2: Calculating Permanent Difference")
print("="*70)

def detect_header_and_columns(df):
    for i in range(len(df)):
        row = df.iloc[i].astype(str).str.upper().tolist()

        if any("PARTICULAR" in c for c in row) and \
           any("DEBIT" in c for c in row) and \
           any("CREDIT" in c for c in row):

            col_map = {}
            for idx, val in enumerate(row):
                if "PARTICULAR" in val:
                    col_map["particulars"] = idx
                elif "DEBIT" in val:
                    col_map["debit"] = idx
                elif "CREDIT" in val:
                    col_map["credit"] = idx

            print(f"  ✅ Header found at Excel Row {i+1}")
            print(f"     Columns → {col_map}")
            return i, col_map

    print("  ❌ Header not found")
    return None, None


def normalize_uid(val: str) -> str:
    if pd.isna(val):
        return ""
    val = str(val).upper().strip()
    while "  " in val:
        val = val.replace("  ", " ")
    return val


def calculate_credit_minus_debit(file_path, label, valid_ids):
    print(f"\n📂 Processing {label}")
    print(f"  Valid UniqueIDs count: {len(valid_ids)}")

    df = pd.read_excel(file_path, header=None)

    # Detect header
    header_idx, cols = detect_header_and_columns(df)
    if header_idx is None:
        print(f"  ⚠️ Skipping {label}")
        return 0.0

    # Load data after header
    data_df = df.iloc[header_idx + 1:].copy()

    # -------------------------------
    # STEP A: Extract UniqueID from column index 2
    # -------------------------------
    data_df["RawUniqueID"] = data_df.iloc[:, 2]
    data_df["NormalizedUniqueID"] = data_df["RawUniqueID"].apply(normalize_uid)

    # -------------------------------
    # STEP B: Normalize valid IDs
    # -------------------------------
    valid_ids_normalized = {normalize_uid(x) for x in valid_ids}

    # -------------------------------
    # STEP C: Filter rows by UniqueID
    # -------------------------------
    before = len(data_df)
    data_df = data_df[
        data_df["NormalizedUniqueID"].isin(valid_ids_normalized)
    ]
    after = len(data_df)

    print(f"  Rows after UniqueID filter: {after} (removed {before-after})")

    # -------------------------------
    # STEP D: Convert Debit / Credit
    # -------------------------------
    data_df["Debit"] = pd.to_numeric(
        data_df.iloc[:, cols["debit"]],
        errors="coerce"
    ).fillna(0)

    data_df["Credit"] = pd.to_numeric(
        data_df.iloc[:, cols["credit"]],
        errors="coerce"
    ).fillna(0)

    # -------------------------------
    # STEP E: Calculate per UniqueID
    # -------------------------------
    grouped = (
        data_df
        .groupby("NormalizedUniqueID")[["Debit", "Credit"]]
        .sum()
    )

    grouped["Diff"] = grouped["Credit"] - grouped["Debit"]

    total_diff = grouped["Diff"].sum()

    print(f"  Unique IDs considered : {len(grouped)}")
    print(f"  Total Debit           : ₹{grouped['Debit'].sum():,.2f}")
    print(f"  Total Credit          : ₹{grouped['Credit'].sum():,.2f}")
    print(f"  Credit - Debit        : ₹{total_diff:,.2f}")

    return total_diff


# -------------------------------
# GST FILES
# -------------------------------
gst_total = 0.0

print("\nProcessing GST Ledgers:")
for rate, file_path in GST_FILES.items():
    if file_path is None:
        continue
    gst_diff = calculate_credit_minus_debit(
        file_path,
        f"{rate}% CGST",
    )
    gst_total += gst_diff

gst_total_times_2 = gst_total * 2

print(f"\n🔁 GST Total (before ×2): ₹{gst_total:,.2f}")
print(f"🔁 GST Total (×2)       : ₹{gst_total_times_2:,.2f}")

# -------------------------------
# REGISTERED AGREEMENT
# -------------------------------
print("\nProcessing Registered Agreement:")
registered_agreement_diff = calculate_credit_minus_debit(
    REG_AGREEMENT_FILE,
    "Registered Agreement",
)

# -------------------------------
# FINAL PERMANENT DIFFERENCE
# -------------------------------
permanent_diff = gst_total_times_2 + registered_agreement_diff

print("\n" + "="*70)
print(f"✅ FINAL PERMANENT DIFFERENCE: ₹{permanent_diff:,.2f}")
print("="*70)

# =========================================================
# STEP 4 – PER UNIQUEID ERP vs TALLY RECONCILIATION
# =========================================================
print("\n" + "="*80)
print("STEP 4: Per-UniqueID Receipt Reconciliation")
print("="*80)

# ---------------------------------------------------------
# STEP 4.1: Aggregate Tally Net Balance per UniqueID
# ---------------------------------------------------------
print("\n[STEP 4.1] Aggregating Tally Net Balance per UniqueID...")

tally_receipts_df = (
    tally_df
    .groupby("NormalizedUniqueID", as_index=False)
    .agg({
        "net_balance": "sum"
    })
    .rename(columns={
        "net_balance": "Tally_Net_Balance"
    })
)

print(f"✓ Tally customers aggregated: {len(tally_receipts_df)}")

# ---------------------------------------------------------
# STEP 4.2: Aggregate ERP Total Received per UniqueID
# ---------------------------------------------------------
print("\n[STEP 4.2] Aggregating ERP Total Received per UniqueID...")

erp_receipts_df = (
    erp_df
    .groupby("NormalizedUniqueID", as_index=False)
    .agg({
        "Total_Received_Amount": "sum"
    })
)

print(f"✓ ERP customers aggregated: {len(erp_receipts_df)}")

# ---------------------------------------------------------
# STEP 4.3: Merge ERP & Tally
# ---------------------------------------------------------
print("\n[STEP 4.3] Merging ERP and Tally data...")

recon_df = pd.merge(
    tally_receipts_df,
    erp_receipts_df,
    on="NormalizedUniqueID",
    how="outer"
)

# Fill missing values
recon_df["Tally_Net_Balance"] = recon_df["Tally_Net_Balance"].fillna(0)
recon_df["Total_Received_Amount"] = recon_df["Total_Received_Amount"].fillna(0)

print(f"✓ Reconciliation rows created: {len(recon_df)}")

# ---------------------------------------------------------
# STEP 4.4: Allocate Permanent Difference per UniqueID
# ---------------------------------------------------------
print("\n[STEP 4.4] Allocating Permanent Difference per customer...")

num_customers = len(recon_df)

if num_customers == 0:
    raise ValueError("❌ No customers found for reconciliation")

permanent_diff_per_customer = permanent_diff / num_customers

print(f"✓ Permanent Difference (Total): ₹{permanent_diff:,.2f}")
print(f"✓ Customers Count              : {num_customers}")
print(f"✓ Per-Customer Adjustment      : ₹{permanent_diff_per_customer:,.2f}")

recon_df["Permanent_Diff_Allocated"] = permanent_diff_per_customer

# ---------------------------------------------------------
# STEP 4.5: Calculate Value of Receipt per UniqueID
# ---------------------------------------------------------
print("\n[STEP 4.5] Calculating Value of Receipt per UniqueID...")

recon_df["Value_Of_Receipt"] = (
    recon_df["Tally_Net_Balance"]
    + recon_df["Permanent_Diff_Allocated"]
)

# ---------------------------------------------------------
# STEP 4.6: Calculate Difference per UniqueID
# ---------------------------------------------------------
print("\n[STEP 4.6] Calculating Difference per UniqueID...")

recon_df["Difference"] = (
    recon_df["Total_Received_Amount"]
    - recon_df["Value_Of_Receipt"]
)

# ---------------------------------------------------------
# STEP 4.7: Debug Output
# ---------------------------------------------------------
print("\n[STEP 4.7] Sample reconciliation output:")
print(
    recon_df[
        [
            "NormalizedUniqueID",
            "Tally_Net_Balance",
            "Permanent_Diff_Allocated",
            "Value_Of_Receipt",
            "Total_Received_Amount",
            "Difference"
        ]
    ].head(10)
)

# ---------------------------------------------------------
# STEP 4.8: Summary
# ---------------------------------------------------------
print("\n[STEP 4.8] Reconciliation Summary:")

matched = recon_df[recon_df["Difference"].abs() <= 1]
mismatch = recon_df[recon_df["Difference"].abs() > 1]

print(f"✓ Matched customers   : {len(matched)}")
print(f"⚠️ Mismatched customers: {len(mismatch)}")

print("\n⚠️ Sample mismatches:")
print(
    mismatch[
        ["NormalizedUniqueID", "Difference"]
    ].head(10)
)

print("\n✅ STEP 4 COMPLETED – PER CUSTOMER RECONCILIATION DONE")


# =========================================================
# STEP 5 – CREATE FINAL CSV
# =========================================================
print("\n" + "="*80)
print("STEP 5: Creating final CSV output")
print("="*80)

# Map columns from `recon_df` to final output columns
# Some columns like 'additional_permanent_difference' or 'total_permanent_difference' 
# may not exist yet, so we can derive them or copy from what we already have.

output_df = recon_df.copy()

# Add additional permanent difference columns (if needed)
# Here, we assume:
# - 'Permanent_Diff_Allocated' → additional_permanent_difference
# - 'Value_Of_Receipt' → value_of_receipts
# - 'Difference' → DIFFERENCE
output_df["additional_permanent_difference"] = output_df["Permanent_Diff_Allocated"]
output_df["total_permanent_difference"] = permanent_diff  # total GST + registered agreement diff
output_df["permanent_difference"] = output_df["Permanent_Diff_Allocated"]
output_df["value_of_receipts"] = output_df["Value_Of_Receipt"]
output_df["DIFFERENCE"] = output_df["Difference"]

# If you want the original Debit/Credit/UniqueID from Tally
# Merge Tally Debit/Credit to final output
tally_subset = tally_df[["NormalizedUniqueID", "Debit", "Credit", "net_balance"]].copy()
tally_subset = tally_subset.rename(columns={"NormalizedUniqueID": "UniqueID"})

final_output = pd.merge(
    output_df,
    tally_subset,
    left_on="NormalizedUniqueID",
    right_on="UniqueID",
    how="left"
)

# Reorder columns exactly as requested
final_output = final_output[[
    "UniqueID",
    "Debit", 
    "Credit",
    "net_balance",
    "additional_permanent_difference",
    "total_permanent_difference",
    "permanent_difference",
    "value_of_receipts",
    "DIFFERENCE"
]]

# Save CSV
output_file_path = BASE_DIR / "reconcilied_output.csv"
final_output.to_csv(output_file_path, index=False)

print(f"\n✅ Final reconciled CSV saved to: {output_file_path}")
print("="*80)