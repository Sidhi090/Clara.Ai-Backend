import pandas as pd
from pathlib import Path
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================
BASE_DIR = Path("IRC_NEW")

ERP_FILE = BASE_DIR / "erp data.xlsx"
TALLY_FILE = BASE_DIR / "tally data.xls"

GST_FILES = {
    "0.5": BASE_DIR / "05 CGST.xls",
    "2.5": BASE_DIR / "25 CGST.xls",
    "9": None,
    "6": None  # provision for future
}

REG_AGREEMENT_FILE = BASE_DIR / "REGISTERED AGREEMENT.xls"

OTHER_DIFFERENCES_FILE = BASE_DIR / "Other permanent difference.xlsx"

# =========================================================
# STEP 0 – LOAD ERP DATA & CREATE UNIQUE IDENTIFIER
# =========================================================
print("="*70)
print("STEP 0: Loading ERP Data & Creating Unique Identifiers")
print("="*70)

raw_df = pd.read_excel(ERP_FILE, header=None)

# Normalize all cells to string for safe scanning
raw_df = raw_df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")

# ------------------------------------------------------
# STEP 2: Define required columns (logical names)
# ------------------------------------------------------
REQUIRED_COLUMNS = {
    "member_name": ["MEMBER NAME"],
    "unit_code": ["UNIT CODE"],
    "total_received": [
        "TOTAL RECEIVED AMOUNT",
    ],
}

# ------------------------------------------------------
# STEP 3: Find the header row dynamically
# ------------------------------------------------------
print("\n[STEP 3] Scanning rows to detect header...")

header_row_index = None
header_map = {}

for row_idx in range(len(raw_df)):
    row_values = raw_df.iloc[row_idx].str.upper().tolist()

    print(f"\n🔍 Scanning Excel Row {row_idx + 1}")

    if row_idx == 10:
        break
    print(f"Row values: {row_values}")

    temp_map = {}

    for logical_name, possible_names in REQUIRED_COLUMNS.items():
        for col_idx, cell in enumerate(row_values):
            for name in possible_names:
                if name in cell and cell != "":
                    temp_map[logical_name] = col_idx
                    print(
                        f"   ✅ Found '{name}' for '{logical_name}' "
                        f"at column index {col_idx}"
                    )

    print(f"   ➜ Matched columns so far: {temp_map}")

    if len(temp_map) == len(REQUIRED_COLUMNS):
        header_row_index = row_idx
        header_map = temp_map
        print(f"\n🎯 HEADER ROW CONFIRMED at Excel row {row_idx + 1}")
        break

if header_row_index is None:
    raise ValueError("❌ Header row not found. Required columns missing.")

# ------------------------------------------------------
# STEP 4: Load ERP using detected header
# ------------------------------------------------------
print("\n[STEP 4] Reloading Excel using detected header row...")
erp_df = pd.read_excel(ERP_FILE, header=header_row_index)

print(f"ERP DataFrame shape: {erp_df.shape}")
print("Detected column names:")
for i, col in enumerate(erp_df.columns):
    print(f"  {i}: {col}")

# Clean column names
erp_df.columns = erp_df.columns.astype(str).str.strip()

# Drop empty rows
before_rows = len(erp_df)
erp_df = erp_df.dropna(how="all")
after_rows = len(erp_df)

print(f"Dropped {before_rows - after_rows} empty rows")

# ------------------------------------------------------
# STEP 5: Resolve actual column names
# ------------------------------------------------------
member_name_col = erp_df.columns[header_map["member_name"]]
unit_code_col = erp_df.columns[header_map["unit_code"]]
total_received_col = erp_df.columns[header_map["total_received"]]

print("\n[STEP 5] Final column mapping:")
print(f"  Member Name Column        → {member_name_col}")
print(f"  Unit Code Column          → {unit_code_col}")
print(f"  Total Received Amount Col → {total_received_col}")

# ------------------------------------------------------
# STEP 6: Create UniqueID
# ------------------------------------------------------
print("\n[STEP 6] Creating UniqueID...")

erp_df["UniqueID"] = (
    erp_df[unit_code_col].astype(str).str.strip()
    + " "
    + erp_df[member_name_col].astype(str).str.strip()
)

# Remove junk rows
before_rows = len(erp_df)
erp_df = erp_df[
    (erp_df[unit_code_col].astype(str).str.strip() != "")
    & (erp_df[member_name_col].astype(str).str.strip() != "")
]
after_rows = len(erp_df)

print(f"Removed {before_rows - after_rows} invalid rows")

print("\nSample rows with UniqueID:")
print(
    erp_df[
        [unit_code_col, member_name_col, total_received_col, "UniqueID"]
    ].head(10)
)

print("\n✅ DONE: ERP data processed successfully")


# Ensure Total Received Amount is numeric
if total_received_col:
    erp_df["Total_Received_Amount"] = pd.to_numeric(
        erp_df[total_received_col].astype(str).str.replace(',', ''), 
        errors='coerce'
    ).fillna(0)
else:
    erp_df["Total_Received_Amount"] = 0

erp_unique_ids = set(erp_df["UniqueID"])

print(f"\n✓ Loaded {len(erp_df)} records from ERP")
print(f"✓ Created {len(erp_unique_ids)} unique identifiers")


# =========================================================
# STEP 1A – LOAD TALLY DATA
# =========================================================
print("\n" + "="*70)
print("STEP 1A: Loading Tally Data")
print("="*70)

print("\n[STEP 1] Loading entire Tally sheet without header...")

raw_df = pd.read_excel(TALLY_FILE, header=None)
raw_df = raw_df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")

print(f"Loaded raw sheet with shape: {raw_df.shape}")

print("\nPreview first 15 rows:")
print(raw_df.head(15))

# ------------------------------------------------------
# STEP 2: Required columns
# ------------------------------------------------------
print("\n[STEP 2] Required columns definition...")

REQUIRED_COLUMNS = {
    "particulars": ["PARTICULARS"],
    "debit": ["DEBIT"],
    "credit": ["CREDIT"],
}

for k, v in REQUIRED_COLUMNS.items():
    print(f"  {k}: {v}")

# ------------------------------------------------------
# STEP 3: Detect header row
# ------------------------------------------------------
print("\n[STEP 3] Detecting MULTI-ROW header...")

header_row_index = None
header_map = {}

WINDOW_SIZE = 4  # how many consecutive rows to inspect

for start_row in range(len(raw_df) - WINDOW_SIZE):
    print(f"\n🔍 Checking rows {start_row + 1} to {start_row + WINDOW_SIZE}")

    combined_cells = {}

    for row_offset in range(WINDOW_SIZE):
        row_idx = start_row + row_offset
        row_values = raw_df.iloc[row_idx].str.upper().tolist()

        print(f"   Row {row_idx + 1}: {row_values}")

        for col_idx, cell in enumerate(row_values):
            if cell == "":
                continue

            # Collect all header words seen per column
            combined_cells.setdefault(col_idx, set()).add(cell)

    temp_map = {}

    for logical_col, possible_names in REQUIRED_COLUMNS.items():
        for col_idx, texts in combined_cells.items():
            for text in texts:
                for name in possible_names:
                    if name in text:
                        temp_map[logical_col] = col_idx
                        print(
                            f"   ✅ Found '{name}' for '{logical_col}' "
                            f"in column {col_idx}"
                        )

    print(f"   ➜ Combined matches: {temp_map}")

    if len(temp_map) == len(REQUIRED_COLUMNS):
        header_row_index = start_row
        header_map = temp_map
        print(
            f"\n🎯 HEADER CONFIRMED using rows "
            f"{start_row + 1} to {start_row + WINDOW_SIZE}"
        )
        break

if header_row_index is None:
    raise ValueError("❌ Header row not detected even with multi-row logic")


# ------------------------------------------------------
# STEP 4: Reload using detected header
# ------------------------------------------------------
print("\n[STEP 4] Reloading Excel using detected header...")

tally_df = pd.read_excel(
    TALLY_FILE,
    header=header_row_index
)

tally_df.columns = tally_df.columns.astype(str).str.strip()
tally_df = tally_df.dropna(how="all")

print(f"Tally DataFrame shape after reload: {tally_df.shape}")

print("\nDetected columns:")
for i, col in enumerate(tally_df.columns):
    print(f"  {i}: {col}")

# ------------------------------------------------------
# STEP 5: Resolve column names
# ------------------------------------------------------
particulars_col = tally_df.columns[header_map["particulars"]]
debit_col = tally_df.columns[header_map["debit"]]
credit_col = tally_df.columns[header_map["credit"]]

print("\n[STEP 5] Column mapping:")
print(f"  Particulars → {particulars_col}")
print(f"  Debit       → {debit_col}")
print(f"  Credit      → {credit_col}")

# ------------------------------------------------------
# STEP 6: Find 'Flats' row and cut data
# ------------------------------------------------------
print("\n[STEP 6] Locating 'Flats' row...")

flats_mask = tally_df[particulars_col].str.upper() == "FLATS"

if not flats_mask.any():
    raise ValueError("❌ 'Flats' row not found — cannot determine data start")

flats_row_index = flats_mask.idxmax()

print(f"✅ 'Flats' found at DataFrame index: {flats_row_index}")

print("\nRows ABOVE Flats (to be removed):")
print(tally_df.loc[:flats_row_index, [particulars_col]].tail(10))

# Keep only rows AFTER Flats
tally_df = tally_df.loc[flats_row_index + 1:].reset_index(drop=True)

print(f"\nRows remaining AFTER Flats: {len(tally_df)}")

# ------------------------------------------------------
# STEP 7: Create UniqueID
# ------------------------------------------------------
print("\n[STEP 7] Creating UniqueID...")

tally_df["UniqueID"] = tally_df[particulars_col].astype(str).str.strip()

# ------------------------------------------------------
# STEP 8: Convert Debit / Credit
# ------------------------------------------------------
print("\n[STEP 8] Converting Debit & Credit to numeric...")

tally_df["Debit"] = pd.to_numeric(
    tally_df[debit_col].astype(str).str.replace(",", ""),
    errors="coerce"
).fillna(0)

tally_df["Credit"] = pd.to_numeric(
    tally_df[credit_col].astype(str).str.replace(",", ""),
    errors="coerce"
).fillna(0)

# ------------------------------------------------------
# STEP 9: Remove junk rows after Flats
# ------------------------------------------------------
print("\n[STEP 9] Removing non-flat rows...")

before = len(tally_df)

tally_df = tally_df[
    tally_df["UniqueID"].str.contains("SM-", case=False, na=False)
]

after = len(tally_df)

print(f"Removed {before - after} non-flat rows")

# ------------------------------------------------------
# STEP 10: Final check
# ------------------------------------------------------
print("\n[STEP 10] Final data preview:")
print(
    tally_df[
        ["UniqueID", "Debit", "Credit"]
    ].head(10)
)

print(f"\n✅ FINAL TALLY RECORD COUNT: {len(tally_df)}")


# =========================================================
# STEP 1B – COMPARE UNIQUE IDs (ERP vs TALLY)
# =========================================================
print("\n" + "=" * 80)
print("STEP 1B: Comparing Unique Identifiers (CLEAN + NORMALIZED)")
print("=" * 80)

# ---------------------------------------------------------
# STEP 1: Normalization function
# ---------------------------------------------------------
def normalize_unique_id(val: str) -> str:
    """
    Normalize UniqueID for reliable comparison
    """
    if val is None:
        return ""

    val = str(val).upper().strip()

    # Remove multiple spaces
    while "  " in val:
        val = val.replace("  ", " ")

    # Remove trailing Dr / Cr
    if val.endswith(" DR"):
        val = val[:-3]
    if val.endswith(" CR"):
        val = val[:-3]

    # Remove pure NAN values
    if val in {"NAN", "NAN NAN"}:
        return ""

    return val


# ---------------------------------------------------------
# STEP 2: Normalize & CLEAN ERP UniqueIDs
# ---------------------------------------------------------
print("\n[STEP 2] Normalizing ERP UniqueIDs...")

erp_df["NormalizedUniqueID"] = erp_df["UniqueID"].apply(normalize_unique_id)

# REMOVE invalid ERP IDs
erp_df = erp_df[
    erp_df["NormalizedUniqueID"].notna()
    & (erp_df["NormalizedUniqueID"] != "")
]

erp_unique_ids = set(erp_df["NormalizedUniqueID"])

print(f"✓ ERP UniqueIDs count (after cleanup): {len(erp_unique_ids)}")

# ---------------------------------------------------------
# STEP 3: Normalize & CLEAN Tally UniqueIDs
# ---------------------------------------------------------
print("\n[STEP 3] Normalizing Tally UniqueIDs...")

tally_df["NormalizedUniqueID"] = tally_df["UniqueID"].apply(normalize_unique_id)

# 🔥 REMOVE invalid Tally IDs
tally_df = tally_df[
    tally_df["NormalizedUniqueID"].notna()
    & (tally_df["NormalizedUniqueID"] != "")
]

tally_unique_ids = set(tally_df["NormalizedUniqueID"])

print(f"✓ Tally UniqueIDs count (after cleanup): {len(tally_unique_ids)}")

# ---------------------------------------------------------
# STEP 4: Compare sets
# ---------------------------------------------------------
print("\n[STEP 4] Comparing ERP vs Tally UniqueIDs...")

in_erp_not_tally = erp_unique_ids - tally_unique_ids
in_tally_not_erp = tally_unique_ids - erp_unique_ids
common_customers = erp_unique_ids & tally_unique_ids

print("\n📊 Comparison Results:")
print(f"  ✓ Common customers              : {len(common_customers)}")
print(f"  ⚠️  In ERP but NOT in Tally     : {len(in_erp_not_tally)}")
print(f"  ⚠️  In Tally but NOT in ERP     : {len(in_tally_not_erp)}")

# ---------------------------------------------------------
# STEP 5: Debug output
# ---------------------------------------------------------
if in_erp_not_tally:
    print("\n⚠️ Customers in ERP but missing in Tally:")
    for customer in sorted(list(in_erp_not_tally))[:10]:
        print(f"  - {customer}")
    if len(in_erp_not_tally) > 10:
        print(f"  ... and {len(in_erp_not_tally) - 10} more")

if in_tally_not_erp:
    print("\n⚠️ Customers in Tally but missing in ERP:")
    for customer in sorted(list(in_tally_not_erp))[:10]:
        print(f"  - {customer}")
    if len(in_tally_not_erp) > 10:
        print(f"  ... and {len(in_tally_not_erp) - 10} more")

# ---------------------------------------------------------
# STEP 6: Sanity check
# ---------------------------------------------------------
print("\n[STEP 6] Sample matched customers:")
for uid in list(common_customers)[:5]:
    print(f"  ✓ MATCHED: {uid}")

print("\n✅ STEP 1B COMPLETED – NAN ISSUE FIXED")


# =========================================================
# STEP 1C – NET BALANCE CALCULATION
# =========================================================
print("\n" + "="*70)
print("STEP 1C: Calculating Net Balance (Credit - Debit)")
print("="*70)

# Net Balance = Credit - Debit (shows as negative Dr)
tally_df["net_balance"] = tally_df["Credit"] - tally_df["Debit"]

print(f"\n✓ Net Balance calculated for {len(tally_df)} records")

# =========================================================
# STEP 1D – FORM FINAL DATAFRAME & EXPORT CSV
# =========================================================

final_df = pd.merge(
    erp_df[
        ["NormalizedUniqueID", "Total_Received_Amount"]
    ],
    tally_df[
        ["NormalizedUniqueID", "net_balance"]
    ],
    on="NormalizedUniqueID",
    how="outer"
)

final_df.rename(
    columns={
        "net_balance": "Tally_Net_Balance"
    },
    inplace=True
)

final_df["Present_in_ERP"] = final_df["Total_Received_Amount"].notna()
final_df["Present_in_Tally"] = final_df["Tally_Net_Balance"].notna()

final_df["Total_Received_Amount"] = final_df["Total_Received_Amount"].fillna(0)
final_df["Tally_Net_Balance"] = final_df["Tally_Net_Balance"].fillna(0)

final_df = final_df[
    [
        "NormalizedUniqueID",
        "Tally_Net_Balance",
        "Total_Received_Amount",
        "Present_in_ERP",
        "Present_in_Tally",
    ]
]

def normalize_uid(val):
    if pd.isna(val):
        return ""
    val = str(val).upper().strip()
    while "  " in val:
        val = val.replace("  ", " ")
    return val


def detect_header_and_columns(df):
    for i in range(len(df)):
        row = df.iloc[i].astype(str).str.upper().tolist()

        if (
            any("PARTICULAR" in c for c in row)
            and any("DEBIT" in c for c in row)
            and any("CREDIT" in c for c in row)
        ):
            col_map = {}
            for idx, val in enumerate(row):
                if "PARTICULAR" in val:
                    col_map["particulars"] = idx
                elif "DEBIT" in val:
                    col_map["debit"] = idx
                elif "CREDIT" in val:
                    col_map["credit"] = idx

            return i, col_map

    return None, None

def calculate_uid_diff_from_file(file_path, target_uid):
    print("\n" + "-" * 60)
    print(f"📂 Loading file: {file_path}")
    print(f"🎯 Target UID  : {target_uid}")

    if file_path is None:
        print("⚠️ File path is None → skipping")
        return 0.0

    df = pd.read_excel(file_path, header=None)
    print(f"✓ Raw file loaded | Shape: {df.shape}")

    print("\n🔍 Detecting header...")
    header_idx, cols = detect_header_and_columns(df)

    if header_idx is None:
        print("❌ Header not detected → skipping file")
        return 0.0

    print(f"✓ Header found at Excel row: {header_idx + 1}")
    print(f"✓ Column map: {cols}")

    data_df = df.iloc[header_idx + 1:].copy()
    print(f"✓ Data rows after header: {len(data_df)}")

    # -------------------------------
    # UniqueID extraction
    # -------------------------------
    print("\n🔎 Normalizing UniqueIDs (column index 2)...")
    data_df["NormalizedUniqueID"] = data_df.iloc[:, 2].apply(normalize_uid)

    print("Sample Normalized UIDs:")
    print(data_df["NormalizedUniqueID"].head(5).tolist())

    # -------------------------------
    # Filter by UID
    # -------------------------------
    before = len(data_df)
    data_df = data_df[data_df["NormalizedUniqueID"] == target_uid]
    after = len(data_df)

    print(f"🧹 Rows matching UID: {after} (removed {before - after})")

    if data_df.empty:
        print("⚠️ No matching rows found for this UID")
        return 0.0

    # -------------------------------
    # Convert Debit / Credit
    # -------------------------------
    print("\n💰 Converting Debit & Credit to numeric...")

    data_df["Debit"] = pd.to_numeric(
        data_df.iloc[:, cols["debit"]],
        errors="coerce"
    ).fillna(0)

    data_df["Credit"] = pd.to_numeric(
        data_df.iloc[:, cols["credit"]],
        errors="coerce"
    ).fillna(0)

    print("\n📄 Sample rows used for calculation:")
    print(
        data_df[
            ["NormalizedUniqueID", "Debit", "Credit"]
        ].head(5)
    )

    diff = (data_df["Credit"] - data_df["Debit"]).sum()

    print(f"✅ Credit Sum : ₹{data_df['Credit'].sum():,.2f}")
    print(f"✅ Debit Sum  : ₹{data_df['Debit'].sum():,.2f}")
    print(f"➡️ Diff (C-D) : ₹{diff:,.2f}")

    return diff


permanent_diff_values = []

for uid in final_df["NormalizedUniqueID"]:
    uid = normalize_uid(uid)

    print("\n" + "=" * 80)
    print(f"🧮 Calculating Permanent Difference for UID: {uid}")
    print("=" * 80)

    # ---------------- GST ----------------
    gst_total = 0.0

    for rate, gst_file in GST_FILES.items():
        print(f"\n➡️ Processing GST {rate}%")

        if gst_file is None:
            print("⚠️ GST file is None → skipped")
            continue

        gst_diff = calculate_uid_diff_from_file(
            gst_file,
            uid,
        )

        print(f"✔ GST {rate}% Diff: ₹{gst_diff:,.2f}")
        gst_total += gst_diff

    gst_total *= 2  # CGST + SGST
    print(f"\n🔁 GST Total after ×2: ₹{gst_total:,.2f}")

    # ---------- REGISTERED AGREEMENT ----------
    print("\n➡️ Processing Registered Agreement")

    reg_diff = calculate_uid_diff_from_file(
        REG_AGREEMENT_FILE,
        uid,
    )

    print(f"✔ Registered Agreement Diff: ₹{reg_diff:,.2f}")

    total_perm_diff = gst_total + reg_diff
    print(f"\n🎯 FINAL Permanent Difference for UID: ₹{total_perm_diff:,.2f}")

    permanent_diff_values.append(total_perm_diff)



# =========================================================
# ADD TO FINAL_DF
# =========================================================

final_df.insert(
    loc=final_df.columns.get_loc("Present_in_Tally") + 1,
    column="Permanent_Difference",
    value=permanent_diff_values,
)

print(final_df.head(10))

final_df["Value_of_Receipts"] = (
    final_df["Permanent_Difference"] +
    final_df["Tally_Net_Balance"]
)

print("\n➕ Adding Other Permanent Differences...")

other_df = pd.read_excel(OTHER_DIFFERENCES_FILE)

# Clean columns
other_df.columns = other_df.columns.astype(str).str.strip()

# Normalize UID from Particulars column
other_df["NormalizedUniqueID"] = other_df["Particulars"].apply(normalize_uid)

# Ensure Amount numeric
other_df["Other_Amount"] = pd.to_numeric(
    other_df["Amount"].astype(str).str.replace(",", ""),
    errors="coerce"
).fillna(0)

other_map = (
    other_df.groupby("NormalizedUniqueID")["Other_Amount"]
    .sum()
    .to_dict()
)

# Map only matching UIDs into final_df
final_df["Other_Permanent_Diff"] = (
    final_df["NormalizedUniqueID"]
    .map(other_map)
    .fillna(0)
)

# Add to Value of Receipts (do NOT overwrite)
final_df["Value_of_Receipts"] = (
    final_df["Value_of_Receipts"] +
    final_df["Other_Permanent_Diff"]
)

print("✅ Other permanent difference receipts merged successfully")


final_df.insert(
    loc=final_df.columns.get_loc("Value_of_Receipts") + 1,
    column="Difference",
    value=final_df["Total_Received_Amount"] - final_df["Value_of_Receipts"]
)

print(final_df.head(10))

output_csv_path = BASE_DIR / "final_reconciliation_new.csv"
final_df.to_csv(output_csv_path, index=False)

print(f"\n✅ Final reconciliation CSV saved at: {output_csv_path}")

